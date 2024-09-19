#!/usr/bin/env python3
import openpyxl as xl
import PySimpleGUI as sg

# Download a blocked keywords file from Connect and use in the corresponding field in the application.
# Copy any products description or title into the first field and the app will check for blocked keywords.


def process(text: str, filepath: str):
    wb = xl.load_workbook(filepath)
    sheet = wb.active
    blocked_words = []
    blocks_in_text = []

    # Iterate through all rows of the blocked words document and extract the blocker words
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        for cell in row:
            blocked_words.append(cell.value)

    # check the input text for occurrences of any blocked word
    for word in blocked_words:
        if word in text:
            blocks_in_text.append(word)

    # return blocked words found in the input text as a LiteralString
    output_text = '\n'.join(blocks_in_text)
    return output_text


def blocked_word_main():
    layout = [
        [sg.Text('Select blocked keywords file (download one from Connect):'), sg.Input(), sg.FileBrowse(key='-FILE-')],
        [sg.Text('Input text', size=(20, 1)), sg.Text('                                                                          '),
         sg.Text('Blocked words in input text:', size=(20, 1))],
        [sg.Multiline(size=(65, 20), key='-INPUT-'), sg.Multiline(size=(65, 20), key='-OUTPUT-', disabled=True)],
        [sg.Button('Process text'), sg.Button('Exit')]
    ]
    window = sg.Window('Find blocked keywords', layout)

    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == 'Exit':
            break
        if event == 'Process text':
            if not values['-FILE-']:
                sg.Popup('No file selected. Download a blocked keywords file from Connect and open it with the browse button')
                continue
            words = process(values['-INPUT-'], values['-FILE-'])
            if words == '':
                window['-OUTPUT-'].update('No blocked words')
            else:
                window['-OUTPUT-'].update(words)


blocked_word_main()
