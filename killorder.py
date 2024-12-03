import requests
from tkinter import filedialog
import openpyxl as xl
import time
import PySimpleGUI as sg

# Takes an order report and an API key and kills all orders in the report,
# doesn't matter if they are Pending, Picked or Invoiced.
# Able to fetch orders from API instead, but this will not cancel Picked or Invoiced orders, only Pending. !! THIS WAS NEVER FINISHED !!


def kill_order_main():
    headers = {
        "Authorization": f"api {get_api_key()}",
        "Content-Type": "application/json"
    }

    order_list_to_kill = parse_orders_to_kill()
    orders_to_cancel = []
    orders_to_return = []
    orders_picked = []
    problem_orders = []

    for order in order_list_to_kill:
        json_response = fetch_order_from_api(order)

        if json_response is None:
            problem_orders.append(order)
            continue

        if json_response["OrderDetails"]["OrderRows"][0].get("PickedQuantity") is not None:
            orders_picked.append(json_response["OrderDetails"]["OrderId"])
            continue

        if json_response["OrderDetails"].get("State") == "Pending":
            orders_to_cancel.append(create_cancel_payload(json_response))
        elif json_response["OrderDetails"].get("State") == "Invoiced":
            orders_to_return.append(create_return_payload(json_response))
        else:
            print(f"Order {order} did not have state Pending or Invoiced, skipping order.")
            print(f"State was: {json_response['OrderDetails'].get('State')}")
            continue

    if len(orders_to_cancel) > 0:
        print(f"Amount of orders to cancel: {len(orders_to_cancel)}")
        input("Press ENTER to continue")
        send_cancel_request(orders_to_cancel, headers)

    if len(orders_to_return) > 0:
        print(f"Amount of orders to return: {len(orders_to_return)}")
        input("Press ENTER to continue")
        send_return_request(orders_to_return, headers)

    if len(orders_picked) > 0:
        print(f"Some orders had \"Picked\" status. Amount: {len(orders_picked)}")
        print("Must mark these as delivered before marking as returned (cancel not possible with this status)")
        input("Press ENTER to continue")
        handle_picked_orders(orders_picked, headers)

    if len(problem_orders) > 0:
        print("\nThere were issues with some orders, they were not found when fetching")
        print(problem_orders)


def parse_orders_to_kill():
    button_clicked = open_popup()
    if button_clicked == 'order_report':
        order_excel = xl.load_workbook(filedialog.askopenfilename(initialdir="C:\\Users\\victrosb\\Downloads"))
        ws = order_excel.worksheets[0]
        return [ws.cell(row, 1).value for row in range(2, we.max_row + 1)]

    elif button_clicked == 'fetch_api':
        orders_json = get_orderlist_from_api()
        return [order["OrderDetails"]["OrderId"] for order in orders_json]


def get_api_key():
    while True:
        key = sg.popup_get_text("Please enter the API KEY", "Getting API KEY", default_text="")
        if len(key) != 36:
            sg.popup_ok("API key incorrect format", title="Incorrect format")
        else:
            return key


def get_orderlist_from_api():
    api_key = get_api_key()
    h = {
        "Authorization": f"api {api_key}",
        "Content-Type": "application/json"
    }
    response = requests.get(url="https://admin.marketplace.cdon.com/api/order", headers=h)
    json_response = response.json()
    return json_response


def open_popup():
    layout = [
        [sg.Text('Click a button')],
        [sg.Button('Get orders from Excel file', key='order_report'),
         sg.Button('Get orders from API', key='fetch_api')]
    ]

    window = sg.Window('Popup Window', layout, modal=True)

    while True:
        event, _ = window.read()
        if event in (sg.WINDOW_CLOSED, 'order_report', 'fetch_api'):
            break

    window.close()
    return event


def fetch_order_from_api(order):
    api_key = get_api_key()
    h = {
        "Authorization": f"api {api_key}",
        "Content-Type": "application/json"
    }
    response = requests.get(url=f"https://admin.marketplace.cdon.com/api/order/{order}", headers=h)
    if response.status_code != 404:
        json_response = response.json()
        return json_response
    return None


def create_cancel_payload(json_data):
    payload_bit = initialize_payload_bit(json_data["OrderDetails"].get("OrderId"), "Rows")
    for row in json_data["OrderDetails"]["OrderRows"]:
        payload_bit["Rows"].append(get_rows(row, "Cancel"))
    return payload_bit


def create_return_payload(json_data):
    payload_bit = initialize_payload_bit(json_data["OrderDetails"].get("OrderId"), "Products")
    for row in json_data["OrderDetails"]["OrderRows"]:
        payload_bit["Products"].append(get_rows(row, "Return"))
    return payload_bit


def create_deliver_payload(json_data):
    payload_bit = initialize_payload_bit(json_data["OrderDetails"].get("OrderId"), "Products")
    for row in json_data["OrderDetails"]["OrderRows"]:
        payload_bit["Products"].append(get_rows(row, "Deliver"))
    return payload_bit


# More dynamic function, rather than two separate.
def initialize_payload_bit(order, method):
    return {"OrderId": f"{order}", f"{method}": []}


# def initialize_payload_bit_cancel(order):
#     return {"OrderId": f"{order}", "Rows": []}
#
#
# def initialize_payload_bit_return(order):
#     return {"OrderId": f"{order}", "Products": []}


# More dynamic than the three previous get_rows functions.
def get_rows(row, method):
    return {"OrderRowId": row["OrderRowId"], f"QuantityTo{method}": row["Quantity"]}


# def get_cancel_rows(row):
#     return {"OrderRowId": row.get("OrderRowId"), "QuantityToCancel": row.get("Quantity")}
#
#
# def get_return_rows(row):
#     return {"OrderRowId": row.get("OrderRowId"), "QuantityToReturn": row.get("Quantity")}
#
#
# def get_deliver_rows(row):
#     return {"OrderRowId": row.get("OrderRowId"), "QuantityToDeliver": row.get("Quantity")}


def send_cancel_request(payload_list, h):
    i = 0
    for bit in payload_list:
        response = requests.post(url="https://admin.marketplace.cdon.com/api/ordercancel",
                                 json=bit, headers=h)
        print(f"\nSent cancel request for order {bit.get('OrderId')}")

        if response.status_code != 200:
            print(f"Request was unsuccessful, status code: {response.status_code}")
            print(response.content)
        else:
            print(f"Request was successful, status code: {response.status_code}")
        print(f"Remaining products: {len(payload_list) - i} / {len(payload_list)}")
        i += 1


def send_return_request(payload_list, h):
    i = 0
    for bit in payload_list:
        response = requests.post(url="https://admin.marketplace.cdon.com/api/orderreturn",
                                 json=bit, headers=h)
        print(f"\nSent return request for order {bit.get('OrderId')}")

        if response.status_code != 200:
            print(f"Request was unsuccessful, status code: {response.status_code}")
            print(response.content)
        else:
            print(f"Request was successful, status code: {response.status_code}")
        print(f"Remaining products: {len(payload_list) - i} / {len(payload_list)}")
        i += 1


def send_deliver_request(payload, h):
    for bit in payload:
        response = requests.post(url="https://admin.marketplace.cdon.com/api/orderdelivery",
                                 json=bit, headers=h)
        print(f"\nSent delivery request for order {bit.get('OrderId')}")

        if response.status_code != 200:
            print(f"Request was unsuccessful, status code: {response.status_code}")
            print(response.content)
        else:
            print(f"Request was successful, status code: {response.status_code}")


def handle_picked_orders(orders, h):
    orders_to_return_json_list = []
    orders_to_deliver_json_list = []

    for order in orders:
        json_response = fetch_order_from_api(order)
        orders_to_deliver_json_list.append(create_deliver_payload(json_response))
        orders_to_return_json_list.append(create_return_payload(json_response))

    send_deliver_request(orders_to_deliver_json_list, h)
    time.sleep(0.1)
    send_return_request(orders_to_return_json_list, h)
