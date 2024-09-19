import openpyxl as xl
from tkinter import filedialog
import xml.etree.ElementTree as et
import json


def idas_main():
    print("Open the Excel file with product SKUs")
    product_excel = xl.load_workbook(filedialog.askopenfilename(initialdir="C:\\Users\\victrosb\\Downloads"))
    ws = product_excel.worksheets[0]

    sku_list = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]

    print("Open the merchant JSON inventory report.")
    json_inventory = open(filedialog.askopenfilename(initialdir="C:\\Users\\victrosb\\Downloads"), 'r', encoding="utf-8")
    json_content = json.load(json_inventory)
    json_inventory.close()

    namespace = "https://schemas.cdon.com/product/4.0/4.12.1/availability"
    marketplace = et.Element("marketplace", xmlns=namespace)
    i = 0
    for product in json_content["Products"]:
        if product["SKU"] not in sku_list:
            continue
        i += 1
        product_element = et.SubElement(marketplace, "product")
        et.SubElement(product_element, "id").text = str(product["SKU"])
        et.SubElement(product_element, "stock").text = "100"

        if product["StatusSe"] != "":
            se = et.SubElement(product_element, "se")
            et.SubElement(se, "status").text = "Online"
            se_delivery_time = et.SubElement(se, "deliveryTime")
            et.SubElement(se_delivery_time, "min").text = "3"
            et.SubElement(se_delivery_time, "max").text = "6"
        if product["StatusDk"] != "":
            dk = et.SubElement(product_element, "dk")
            et.SubElement(dk, "status").text = "Online"
            dk_delivery_time = et.SubElement(dk, "deliveryTime")
            et.SubElement(dk_delivery_time, "min").text = "3"
            et.SubElement(dk_delivery_time, "max").text = "6"

        if product["StatusNo"] != "":
            no = et.SubElement(product_element, "no")
            et.SubElement(no, "status").text = "Online"
            no_delivery_time = et.SubElement(no, "deliveryTime")
            et.SubElement(no_delivery_time, "min").text = "3"
            et.SubElement(no_delivery_time, "max").text = "6"

        if product["StatusFi"] != "":
            fi = et.SubElement(product_element, "fi")
            et.SubElement(fi, "status").text = "Online"
            fi_delivery_time = et.SubElement(fi, "deliveryTime")
            et.SubElement(fi_delivery_time, "min").text = "3"
            et.SubElement(fi_delivery_time, "max").text = "6"
    print(f"Will save {i} amount of products in XML file")
    filename = input("Enter the filename to save: ")
    if not filename.endswith(".xml"):
        filename += ".xml"

    tree = et.ElementTree(marketplace)
    tree.write(filename, encoding="utf-8", xml_declaration=True)

idas_main()