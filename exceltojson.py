import openpyxl as xl
import json
from tkinter import filedialog
import os
from datetime import datetime


def excel_to_json_main():
    workbook = open_workbook()
    worksheet = workbook.worksheets[0]

    header = get_header(worksheet)
    values = get_values(worksheet)
    zipped_dict = zip_lists(header, values)
    finished_dict = create_json_data(zipped_dict)

    json_data = json.dumps(finished_dict, indent=4)
    write_file(json_data)


def open_workbook():
    cwd = os.getcwd()
    return xl.load_workbook(filedialog.askopenfilename(initialdir=cwd))


def get_header(worksheet):
    header = []
    for column in range(1, worksheet.max_column + 1):
        position = worksheet.cell(1, column)
        header.append(position.value)
    return header


def get_values(worksheet):
    value_list = []

    for row in range(2, worksheet.max_row + 1):
        temp_value_list = []
        for column in range(1, worksheet.max_column + 1):
            position = worksheet.cell(row, column)
            temp_value_list.append(position.value)
        value_list.append(temp_value_list)
    return value_list


def zip_lists(header, values):
    finished_list = []
    for i in range(len(values)):
        finished_list.append(dict(zip(header, values[i])))
    return finished_list


def create_json_data(dict_list):
    products = {"products": []}
    for product in dict_list:
        products["products"].append(product)
    return products


def write_file(data):
    date = datetime.now().strftime("%Y_%m_%d-%H-%M-%S_%f")
    with open(f"json_result_{date}.json", "w", encoding="utf-8") as json_file:
        json_file.write(data)
