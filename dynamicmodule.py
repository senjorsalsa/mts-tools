import json
import os
import xml.etree.ElementTree as et
import PySimpleGUI as sg
import openpyxl as xl


def dynamic_main():
    excel_filepath, json_filepath = get_filepaths()
    json_data = get_json_data(json_filepath)
    workbook = xl.load_workbook(excel_filepath)
    products = get_products(workbook)
    cdon_id_list = get_cdon_ids(products, json_data)
    with open("product ids.txt", "w") as f:
        for product_id in cdon_id_list:
            f.write(product_id + "\n")


def get_products(workbook):
    active_sheet = workbook.active
    products = []

    for row in range(2, active_sheet.max_row + 1):
        column = 1
        position = active_sheet.cell(row, column)
        products.append(position.value)
    return products


def get_filepaths():
    layout = [
        [sg.Input(), sg.FileBrowse("Choose Excel file", key="filepath_excel")],
        [sg.Input(), sg.FileBrowse("Choose JSON file", key="filepath_json")],
        [sg.Button("Accept", key="OK")]
    ]
    window = sg.Window("Select filepaths", layout)
    event, values = window.read()

    while True:
        if event == sg.WINDOW_CLOSED:
            break
        elif event == "OK":
            excel_filepath = values["filepath_excel"]
            json_filepath = values["filepath_json"]
            return excel_filepath, json_filepath


def get_json_data(filepath):
    with open(filepath, encoding="utf-8") as json_raw:
        json_data = json.load(json_raw)
    return json_data


def get_cdon_ids(product_list, json_data):
    return [product["CdonProductId"] for product in json_data["Products"] if product["SKU"] in product_list]
